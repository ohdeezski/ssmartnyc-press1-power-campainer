import csv
import os
import re

import phonenumbers
from openpyxl import load_workbook
from flask import current_app

from app.extensions import db
from app.modules.contacts.models import Contact, ContactList


E164_RE = re.compile(r"^\+[1-9]\d{1,14}$")


def normalize_phone(phone_str):
    """Parse and return E.164 format, or None if unparseable."""
    if not phone_str or not phone_str.strip():
        return None
    cleaned = phone_str.strip().replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    for region in (None, "US"):
        try:
            parsed = phonenumbers.parse(cleaned, region)
            return phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
        except phonenumbers.NumberParseException:
            continue
    return None


def is_blocked(phone):
    """Check if a number is on the global blocklist."""
    blacklist_path = os.path.join(current_app.instance_path, "numbers", "blacklist.txt")
    if not os.path.exists(blacklist_path):
        return False
    with open(blacklist_path, "r") as f:
        blocked = {line.strip() for line in f if line.strip()}
    return phone in blocked


class ContactImportService:
    @staticmethod
    def parse_file(file_path, file_category):
        """Parse a TXT/CSV/XLSX file and return a list of raw phone strings."""
        raw_numbers = []

        if file_category == "txt":
            with open(file_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#"):
                        raw_numbers.append(line)

        elif file_category == "csv":
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    phone = row.get("phone") or row.get("number") or row.get("mobile") or row.get("tel")
                    if phone:
                        raw_numbers.append(phone.strip())

        elif file_category == "xlsx":
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            phone_col = None
            for idx, h in enumerate(headers):
                if h and h.lower() in ("phone", "number", "mobile", "tel", "cell"):
                    phone_col = idx
                    break
            if phone_col is None:
                phone_col = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[phone_col] if phone_col < len(row) else None
                if val:
                    raw_numbers.append(str(val).strip())
            wb.close()

        return raw_numbers

    @staticmethod
    def process(raw_numbers, contact_list_id, created_by):
        """Process raw numbers: normalize, dedupe, validate, blocklist check.

        Returns a dict with counts and lists of valid/invalid/duplicate/blocked contacts.
        """
        stats = {"loaded": 0, "duplicates": 0, "invalid": 0, "blocked": 0, "remaining": 0}
        valid_contacts = []
        seen = set()

        # Global dedupe: numbers already in this contact list
        existing = (
            db.session.query(Contact.phone)
            .filter(Contact.contact_list_id == contact_list_id)
            .with_entities(Contact.phone)
            .all()
        )
        existing_phones = {p[0] for p in existing}

        for raw in raw_numbers:
            stats["loaded"] += 1
            normalized = normalize_phone(raw)

            if normalized is None:
                stats["invalid"] += 1
                continue

            if normalized in existing_phones or normalized in seen:
                stats["duplicates"] += 1
                continue

            if is_blocked(normalized):
                stats["blocked"] += 1
                continue

            seen.add(normalized)
            valid_contacts.append(
                Contact(
                    contact_list_id=contact_list_id,
                    phone=normalized,
                    status="ready",
                )
            )

        stats["remaining"] = len(valid_contacts)
        return stats, valid_contacts

    @staticmethod
    def commit_contacts(contact_list_id, contacts):
        """Bulk-insert validated contacts into the database."""
        if not contacts:
            return 0
        db.session.bulk_save_objects(contacts)
        db.session.commit()
        return len(contacts)

    @staticmethod
    def create_list(name, source_file_id, created_by):
        """Create a new ContactList record."""
        cl = ContactList(name=name, source_file_id=source_file_id, created_by=created_by)
        db.session.add(cl)
        db.session.commit()
        return cl
